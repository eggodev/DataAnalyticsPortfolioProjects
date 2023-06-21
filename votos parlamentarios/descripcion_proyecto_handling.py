# Este script de scraping extrae la descripción de los proyectos de ley y los concatena con el detalle de la votación

import openpyxl
import requests
from bs4 import BeautifulSoup

# Leer el archivo de Excel
workbook = openpyxl.load_workbook('buscardesc.xlsx')
sheet = workbook.active

# Crear un diccionario para almacenar los resultados
resultados = {}

# Recorrer la primera columna de la planilla
for row in sheet.iter_rows(min_row=2, max_col=1, values_only=True):
    idVotacion = row[0]

    # Construir la URL con el idVotacion
    url = f"http://silpy.congreso.gov.py/votacion/{idVotacion}"

    # Hacer la solicitud HTTP
    response = requests.get(url)
    soup = BeautifulSoup(response.text, 'html.parser')

    # Extraer el texto de los spans y agregar salto de línea
    spans = soup.find_all('span', class_='itemResaltado1')
    if len(spans) >= 2:
        texto1 = spans[0].text.strip()
        texto2 = spans[1].text.strip()
        resultado = f'{texto1}\nDetalle de la votación: {texto2}'
        resultados[idVotacion] = resultado

# Crear un nuevo archivo de Excel y llenarlo con los resultados
resultados_workbook = openpyxl.Workbook()
resultados_sheet = resultados_workbook.active

for i, idVotacion in enumerate(resultados.keys(), start=1):
    resultado = resultados[idVotacion]
    resultados_sheet.cell(row=i, column=1, value=idVotacion)
    resultados_sheet.cell(row=i, column=2, value=resultado)

# Guardar el archivo de resultados
resultados_workbook.save('resultados.xlsx')
