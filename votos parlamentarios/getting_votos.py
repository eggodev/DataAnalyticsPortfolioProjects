# Este script utilicé para obtener los votos de los Parlamentarios en cada votación. URL: http://silpy.congreso.gov.py

import openpyxl
from bs4 import BeautifulSoup
from urllib.parse import urlparse
import requests
import xlrd

# Leer el archivo Excel "Votaciones.xls"
workbook = xlrd.open_workbook("Votaciones.xls")
sheet = workbook.sheet_by_index(0)

# Obtener los valores de la columna "appURL"
urls = sheet.col_values(1)[1:]  # Excluir el encabezado de la columna

# Crear el libro de Excel y la hoja de trabajo
workbook_resultados = openpyxl.Workbook()
sheet_resultados = workbook_resultados.active

# Modificar las cabeceras de la hoja de trabajo
sheet_resultados["A1"] = "idVotacion"
sheet_resultados["B1"] = "opciones"
sheet_resultados["C1"] = "votos"

# Definir la función para realizar el scraping
def scrape_url(url):
    response = requests.get(url)
    html_content = response.text
    soup = BeautifulSoup(html_content, "html.parser")
    
    tbody_element = soup.find("tbody", id="j_idt97_data")

    votos = []
    ids = []  # Array para almacenar los IDs de los enlaces

    if tbody_element:
        tr_elements = tbody_element.find_all("tr")

        for tr_element in tr_elements:
            if "ui-rowgroup-header" in tr_element.get("class", []):
                if ids:
                    votos.append({span_text: ids})  # Agregar los IDs al array votos

                ids = []  # Reiniciar el array ids
                span_element = tr_element.find("span", class_="font3x")
                if span_element:
                    span_text = span_element.text.strip()  # Obtener el texto del span
                else:
                    span_text = ""  # Si no se encuentra el span, asignar una cadena vacía
            else:
                a_element = tr_element.find("a")
                if a_element:
                    href = a_element.get("href")
                    id_number = href.split("/")[-1]
                    ids.append(id_number)

        # Agregar los últimos IDs al array votos
        if ids:
            votos.append({span_text: ids})

    return votos

# Realizar el scraping para cada URL y agregar los resultados a la planilla de resultados
row = 2
for url in urls:
    # Realizar el scraping para la URL actual
    votos = scrape_url(url)

    parsed_url = urlparse(url)
    url_id = parsed_url.path.split("/")[-1]  # Última sección de la URL como ID

    for voto in votos:
        for index, ids in voto.items():
            sheet_resultados.cell(row=row, column=1, value=url_id)
            sheet_resultados.cell(row=row, column=2, value=index)

            ids_str = ", ".join(ids)  # Convertir los IDs a una cadena separada por comas
            sheet_resultados.cell(row=row, column=3, value=ids_str)
            row += 1

# Guardar el archivo de Excel con los resultados actualizados
workbook_resultados.save("resultados.xlsx")
print("Los datos se han actualizado exitosamente en resultados.xlsx.")
